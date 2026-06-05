import re
import sys
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding='utf-8')
ROOT = Path(__file__).resolve().parent.parent


def slug(name: str) -> str:
    s = name.strip().lower()
    s = re.sub(r'[^a-z0-9]+', '-', s)
    return re.sub(r'-+', '-', s).strip('-')


def count_path_roles(paths_text: str, eid: str) -> int:
    start = paths_text.find(f'expertiseId: "{eid}"')
    if start == -1:
        return 0
    chunk = paths_text[start : start + 12000]
    roles_start = chunk.find('roles: [')
    conn_start = chunk.find('connections: [')
    return chunk[roles_start:conn_start].count("track: '")


title_wb = openpyxl.load_workbook(ROOT / 'dataset/Title.xlsx', read_only=True)
title_rows = list(title_wb['Title'].iter_rows(values_only=True))
title_header = title_rows[0]
exp_i = title_header.index('Expertise')
title_i = title_header.index('Title')

titles_by_exp: dict[str, set[str]] = {}
for r in title_rows[1:]:
    if r[exp_i]:
        titles_by_exp.setdefault(str(r[exp_i]).strip(), set()).add(str(r[title_i]).strip())

paths = (ROOT / 'src/app/data/careerPaths.ts').read_text(encoding='utf-8')

exp_wb = openpyxl.load_workbook(ROOT / 'dataset/Expertise.xlsx', read_only=True)
exp_rows = list(exp_wb.active.iter_rows(values_only=True))
exp_header = [str(c).strip() if c else '' for c in exp_rows[0]]

print(f'{"Expertise":22} {"Title.xlsx":>10} {"careerPaths":>12}')
print('-' * 48)
for r in exp_rows[1:]:
    row = {exp_header[i]: r[i] for i in range(len(exp_header))}
    name = str(row['Expertise']).strip()
    eid = slug(name)
    t_count = len(titles_by_exp.get(name, set()))
    p_count = count_path_roles(paths, eid)
    flag = '  <-- GAP' if t_count > p_count else ''
    print(f'{name:22} {t_count:10} {p_count:12}{flag}')
